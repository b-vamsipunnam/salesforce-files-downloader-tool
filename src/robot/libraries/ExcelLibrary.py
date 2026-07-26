# -*- coding: utf-8 -*-

from io import BytesIO
from typing import Any, Dict, Iterator, List, Optional, Tuple
import os
import shutil
import tempfile
import warnings

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class SuchIdIsExistException(Exception):
    """Raised when the document with the identifier is already in the cache."""
    pass


class NoSuchIdException(Exception):
    """Raised when accessing an absent document identifier."""
    pass


class NoOpenedDocumentsException(Exception):
    """Raised in the absence of open documents."""
    pass


class ExcelLibrary(object):
    """Library for working with Excel documents.

    This is a compatibility wrapper intended to be used as a drop-in Robot Framework library.

    IMPORTANT BEHAVIOR (for backward-compatibility with your Robot suite):
    - 'Create Excel Document' argument is treated as a FILE PATH (not a doc_id).
      It creates the workbook and binds the current document id to that path.
    - 'Save Excel Document' can be called with filename OR with no args; it will save
      to the current document id if it is a path.
    """

    ROBOT_LIBRARY_SCOPE = "GLOBAL"

    def __init__(self) -> None:
        self._cache: Dict[str, openpyxl.Workbook] = {}
        self._current_id: Optional[str] = None

    # -------------------------
    # Helper utilities
    # -------------------------

    def _ensure_parent_dir(self, file_path: str) -> None:
        """Create parent directory if file_path includes a directory."""
        file_path = str(file_path)
        dir_path = os.path.dirname(file_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)

    def _get_current_workbook(self) -> openpyxl.Workbook:
        """Checks opened document and returns current workbook."""
        if not self._cache or not self._current_id:
            raise NoOpenedDocumentsException("No opened documents in cache.")
        return self._cache[self._current_id]

    def get_sheet(self, sheet_name: str = None) -> Worksheet:
        """Returns a sheet from the current document."""
        workbook = self._get_current_workbook()
        if sheet_name is None:
            return workbook.active
        sheet_name = str(sheet_name)
        return workbook[sheet_name]

    # -------------------------
    # Core keywords
    # -------------------------

    def create_excel_document(self, doc_id: str) -> str:
        """
        Creates new excel document.

        COMPAT MODE:
        - 'doc_id' is treated as a FILE PATH (your suite passes ${CV_File_Name} etc.).
        - Creates workbook on disk immediately.
        - Uses the file path as the cache key and current id.
        """
        file_path = str(doc_id)

        # If you call Create twice with same path, keep the old behavior: raise.
        if file_path in self._cache:
            raise SuchIdIsExistException(f"Document with such id {file_path} is created.")

        self._ensure_parent_dir(file_path)

        workbook = openpyxl.Workbook()
        # Save immediately so subsequent steps expecting file existence won't break
        workbook.save(filename=file_path)

        self._cache[file_path] = workbook
        self._current_id = file_path
        return self._current_id

    def open_excel_document(self, filename: str, doc_id: str = None) -> str:
        """
        Opens xlsx document file.

        Supports BOTH signatures:
        - Open Excel Document | filename=file.xlsx | doc_id=myid |
        - Open Excel Document | filename=file.xlsx |   (doc_id omitted)

        If doc_id is omitted, uses filename as id.
        """
        filename = str(filename)
        use_id = str(doc_id) if doc_id is not None else filename

        if use_id in self._cache:
            raise SuchIdIsExistException(f"Document with such id {use_id} is opened.")

        workbook = openpyxl.load_workbook(filename=filename)
        self._cache[use_id] = workbook
        self._current_id = use_id
        return self._current_id

    def open_excel_document_from_stream(self, stream: bytes, doc_id: str) -> str:
        """Opens xlsx document from stream."""
        doc_id = str(doc_id)
        if doc_id in self._cache:
            raise SuchIdIsExistException(f"Document with such id {doc_id} is opened.")
        workbook = openpyxl.load_workbook(filename=BytesIO(stream))
        self._cache[doc_id] = workbook
        self._current_id = doc_id
        return self._current_id

    def switch_current_excel_document(self, doc_id: str) -> Optional[str]:
        """Switches current excel document."""
        doc_id = str(doc_id)
        if doc_id not in self._cache:
            raise NoSuchIdException(f"Document with such id {doc_id} is not opened yet.")
        old_name = self._current_id
        self._current_id = doc_id
        return old_name

    def close_current_excel_document(self) -> Optional[str]:
        """Close and remove the current document from the cache."""
        close_error = None
        if self._current_id is not None:
            workbook = self._cache.pop(self._current_id, None)
            try:
                if workbook is not None:
                    workbook.close()
            except Exception as error:
                close_error = error
            finally:
                self._current_id = None
        if self._cache:
            self._current_id = list(self._cache.keys())[0]
        if close_error is not None:
            raise RuntimeError(
                "Failed to close the current Excel workbook."
            ) from close_error
        return self._current_id

    def close_all_excel_documents(self) -> None:
        """Close every opened document and clear the cache."""
        workbooks = list(self._cache.values())
        self._cache.clear()
        self._current_id = None
        close_errors = []
        for workbook in workbooks:
            try:
                workbook.close()
            except Exception as error:
                close_errors.append(error)
        if close_errors:
            raise RuntimeError(
                f"Failed to close {len(close_errors)} Excel workbook(s)."
            ) from close_errors[0]

    def save_excel_document(self, filename: str = None) -> None:
        """
        Saves the current document to disk.

        COMPAT MODE:
        - If filename is omitted, saves to current_id when it looks like a path.
        - If filename provided, ensures its parent directory exists.
        """
        workbook = self._get_current_workbook()

        target = filename if filename else self._current_id
        if not target:
            raise NoOpenedDocumentsException("No current document to save.")

        target = str(target)
        self._ensure_parent_dir(target)

        workbook.save(filename=target)

        # If user saved to a new filename, rebind id to that file path
        if filename and self._current_id != target:
            # Move cache entry key to new id
            self._cache[target] = workbook
            if self._current_id in self._cache:
                self._cache.pop(self._current_id, None)
            self._current_id = target

    def get_list_sheet_names(self) -> List[str]:
        """Returns a list of sheet names in the current document."""
        workbook = self._get_current_workbook()
        return workbook.sheetnames

    def make_list_from_excel_sheet(self, sheet: Worksheet) -> list:
        """Making list from Excel sheet."""
        data = []
        for row in sheet.values:
            data.append(row)
        return data

    # -------------------------
    # Read keywords
    # -------------------------

    def read_excel_cell(self, row_num: int, col_num: int, sheet_name: str = None) -> Any:
        """Returns content of a cell."""
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        cell: Cell = sheet.cell(row=row_num, column=col_num)
        return cell.value

    def read_excel_row(
        self, row_num: int, col_offset: int = 0, max_num: int = 0, sheet_name: str = None
    ) -> List[Any]:
        """Returns content of a row from the current sheet of the document."""
        row_num = int(row_num)
        col_offset = int(col_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)

        if max_num <= 0:
            # If not provided, read until last column with values (best-effort)
            max_num = sheet.max_column - col_offset

        row_iter: Iterator[Tuple[Cell]] = sheet.iter_rows(
            min_row=row_num,
            max_row=row_num,
            min_col=1 + col_offset,
            max_col=col_offset + max_num,
        )
        row: Tuple[Cell, ...] = next(row_iter)
        return [cell.value for cell in row]

    def read_excel_column(
        self, col_num: int, row_offset: int = 0, max_num: int = 0, sheet_name: str = None
    ) -> List[Any]:
        """Returns content of a column from the current sheet of the document."""
        col_num = int(col_num)
        row_offset = int(row_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)

        if max_num <= 0:
            max_num = sheet.max_row - row_offset

        row_iter: Iterator[Tuple[Cell, ...]] = sheet.iter_rows(
            min_col=col_num,
            max_col=col_num,
            min_row=1 + row_offset,
            max_row=row_offset + max_num,
        )
        return [row[0].value for row in row_iter]

    # -------------------------
    # Write keywords
    # -------------------------

    def write_excel_cell(self, row_num: int, col_num: int, value: Any, sheet_name: str = None) -> None:
        """Writes value to the cell."""
        row_num = int(row_num)
        col_num = int(col_num)
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row=row_num, column=col_num, value=value)

    def write_excel_row(
        self, row_num: int, row_data: List[Any], col_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes a row to the document."""
        row_num = int(row_num)
        col_offset = int(col_offset)
        sheet = self.get_sheet(sheet_name)
        for col_num in range(len(row_data)):
            sheet.cell(row=row_num, column=col_num + col_offset + 1, value=row_data[col_num])

    def write_excel_rows(
        self, rows_data: List[List[Any]], rows_offset: int = 0, col_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes a list of rows to the document."""
        for row_num, row_data in enumerate(rows_data):
            self.write_excel_row(row_num + int(rows_offset) + 1, row_data, col_offset, sheet_name)

    def write_excel_column(
        self, col_num: int, col_data: List[Any], row_offset: int = 0, sheet_name: str = None
    ) -> None:
        """Writes the data to a column."""
        col_num = int(col_num)
        row_offset = int(row_offset)
        sheet = self.get_sheet(sheet_name)
        for row_num in range(len(col_data)):
            sheet.cell(column=col_num, row=row_num + row_offset + 1, value=col_data[row_num])

    def write_migration_rows_atomically(
        self,
        cv_file_name: Optional[str],
        cv_row: int,
        file_title: str,
        version_data_path: str,
        cdl_file_name: Optional[str],
        cdl_row: int,
        content_links: List[Dict[str, Any]],
        write_content_version: bool = True,
        write_content_document_links: bool = True,
    ) -> None:
        """Stage and commit all migration rows for one document as one operation."""
        targets: List[Tuple[str, int, List[List[Any]]]] = []
        if self._as_bool(write_content_version):
            if not cv_file_name:
                raise ValueError("ContentVersion workbook path is required.")
            targets.append(
                (
                    str(cv_file_name),
                    int(cv_row),
                    [[file_title, version_data_path, version_data_path]],
                )
            )

        if self._as_bool(write_content_document_links):
            if not cdl_file_name:
                raise ValueError("ContentDocumentLink workbook path is required.")
            link_rows = [
                [
                    link["ContentDocumentId"],
                    link["LinkedEntityId"],
                    link["ShareType"],
                    link["Visibility"],
                ]
                for link in content_links
            ]
            targets.append((str(cdl_file_name), int(cdl_row), link_rows))

        staged_files: Dict[str, str] = {}
        backup_files: Dict[str, str] = {}
        committed_paths: List[str] = []
        try:
            for target_path, start_row, rows in targets:
                if not os.path.isfile(target_path):
                    raise FileNotFoundError(target_path)
                target_dir = os.path.dirname(os.path.abspath(target_path))
                staged_fd, staged_path = tempfile.mkstemp(
                    prefix=".salesforce_migration_",
                    suffix=".xlsx",
                    dir=target_dir,
                )
                os.close(staged_fd)
                staged_files[target_path] = staged_path
                shutil.copy2(target_path, staged_path)
                workbook = openpyxl.load_workbook(staged_path)
                try:
                    sheet = workbook.active
                    for row_offset, row_data in enumerate(rows):
                        for col_num, value in enumerate(row_data, start=1):
                            sheet.cell(
                                row=start_row + row_offset,
                                column=col_num,
                                value=value,
                            )
                    workbook.save(staged_path)
                finally:
                    workbook.close()

            for target_path, _, _ in targets:
                target_dir = os.path.dirname(os.path.abspath(target_path))
                target_base_name = os.path.splitext(
                    os.path.basename(target_path)
                )[0]
                backup_fd, backup_path = tempfile.mkstemp(
                    prefix=f".{target_base_name}_rollback_recovery_",
                    suffix=".xlsx",
                    dir=target_dir,
                )
                os.close(backup_fd)
                backup_files[target_path] = backup_path
                shutil.copy2(target_path, backup_path)

            for target_path, _, _ in targets:
                os.replace(staged_files[target_path], target_path)
                staged_files.pop(target_path, None)
                committed_paths.append(target_path)
        except Exception as transaction_error:
            rollback_errors = []
            preserved_backups = []
            for target_path in reversed(committed_paths):
                backup_path = backup_files.get(target_path)
                if backup_path and os.path.exists(backup_path):
                    try:
                        os.replace(backup_path, target_path)
                        backup_files.pop(target_path, None)
                    except Exception as rollback_error:
                        rollback_errors.append((target_path, rollback_error))
                        preserved_backups.append(backup_path)

            cleanup_candidates = [
                *staged_files.values(),
                *(
                    backup_path
                    for backup_path in backup_files.values()
                    if backup_path not in preserved_backups
                ),
            ]
            cleanup_errors = self._remove_temporary_files(cleanup_candidates)

            if rollback_errors:
                recovery_details = ", ".join(preserved_backups)
                message = (
                    "Migration workbook transaction failed and rollback was "
                    f"incomplete. Preserved recovery backup(s): {recovery_details}"
                )
                if cleanup_errors:
                    message += (
                        f". {len(cleanup_errors)} additional temporary file(s) "
                        "could not be removed"
                    )
                raise RuntimeError(message) from transaction_error

            if cleanup_errors:
                transaction_error.add_note(
                    f"{len(cleanup_errors)} transaction temporary file(s) "
                    "could not be removed."
                )
            raise
        else:
            cleanup_errors = self._remove_temporary_files(
                [*staged_files.values(), *backup_files.values()]
            )
            if cleanup_errors:
                warnings.warn(
                    f"{len(cleanup_errors)} migration transaction temporary "
                    "file(s) could not be removed.",
                    RuntimeWarning,
                    stacklevel=2,
                )

    @staticmethod
    def _as_bool(value: Any) -> bool:
        if isinstance(value, str):
            return value.strip().lower() in {"true", "yes", "1"}
        return bool(value)

    @staticmethod
    def _remove_temporary_files(paths: List[str]) -> List[Tuple[str, Exception]]:
        """Remove transaction files without masking the primary operation."""
        errors = []
        for path in paths:
            if not os.path.exists(path):
                continue
            try:
                os.remove(path)
            except Exception as error:
                errors.append((path, error))
        return errors
