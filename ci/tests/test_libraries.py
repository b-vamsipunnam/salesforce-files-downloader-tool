import json
import os
import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

import openpyxl

from src.robot.libraries.ExcelLibrary import ExcelLibrary
from src.robot.libraries.ExecutionReporting import FAILURE_CODES, ExecutionReporting
from src.robot.libraries.SalesforceSupport import SalesforceSupport


class ExcelLibraryTransactionTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_directory.cleanup)
        self.cv_path = os.path.join(self.temp_directory.name, "cv.xlsx")
        self.cdl_path = os.path.join(self.temp_directory.name, "cdl.xlsx")
        for path in (self.cv_path, self.cdl_path):
            workbook = openpyxl.Workbook()
            workbook.save(path)
            workbook.close()
        self.library = ExcelLibrary()
        self.links = [
            {
                "ContentDocumentId": "069000000000001",
                "LinkedEntityId": "001000000000001",
                "ShareType": "V",
                "Visibility": "AllUsers",
            }
        ]

    def write_rows(self):
        self.library.write_migration_rows_atomically(
            self.cv_path,
            2,
            "Title",
            "C:/download/file.bin",
            self.cdl_path,
            2,
            self.links,
            True,
            True,
        )

    def cell_value(self, path, row, column):
        workbook = openpyxl.load_workbook(path)
        try:
            return workbook.active.cell(row, column).value
        finally:
            workbook.close()

    def transaction_files(self):
        return [
            name
            for name in os.listdir(self.temp_directory.name)
            if name.startswith(".salesforce_migration_")
            or "_rollback_recovery_" in name
        ]

    def test_commits_both_workbooks_and_removes_temporary_files(self):
        self.write_rows()

        self.assertEqual(self.cell_value(self.cv_path, 2, 1), "Title")
        self.assertEqual(
            self.cell_value(self.cdl_path, 2, 1),
            "069000000000001",
        )
        self.assertEqual(self.transaction_files(), [])

    def test_staging_failure_leaves_originals_and_removes_temporary_files(self):
        with patch(
            "src.robot.libraries.ExcelLibrary.openpyxl.load_workbook",
            side_effect=OSError("simulated staging failure"),
        ):
            with self.assertRaisesRegex(OSError, "staging failure"):
                self.write_rows()

        self.assertIsNone(self.cell_value(self.cv_path, 2, 1))
        self.assertIsNone(self.cell_value(self.cdl_path, 2, 1))
        self.assertEqual(self.transaction_files(), [])

    def test_second_commit_failure_rolls_back_first_workbook(self):
        original_replace = os.replace

        def fail_second_commit(source, target):
            if target == self.cdl_path and os.path.basename(source).startswith(
                ".salesforce_migration_"
            ):
                raise OSError("simulated second commit failure")
            return original_replace(source, target)

        with patch(
            "src.robot.libraries.ExcelLibrary.os.replace",
            side_effect=fail_second_commit,
        ):
            with self.assertRaisesRegex(OSError, "second commit failure"):
                self.write_rows()

        self.assertIsNone(self.cell_value(self.cv_path, 2, 1))
        self.assertIsNone(self.cell_value(self.cdl_path, 2, 1))
        self.assertEqual(self.transaction_files(), [])

    def test_failed_rollback_preserves_recovery_backup(self):
        original_replace = os.replace

        def fail_commit_and_rollback(source, target):
            source_name = os.path.basename(source)
            if target == self.cdl_path and source_name.startswith(
                ".salesforce_migration_"
            ):
                raise OSError("simulated second commit failure")
            if target == self.cv_path and "_rollback_recovery_" in source_name:
                raise OSError("simulated rollback failure")
            return original_replace(source, target)

        with patch(
            "src.robot.libraries.ExcelLibrary.os.replace",
            side_effect=fail_commit_and_rollback,
        ):
            with self.assertRaisesRegex(
                RuntimeError,
                "Preserved recovery backup",
            ):
                self.write_rows()

        recovery_files = [
            name for name in self.transaction_files() if "_rollback_recovery_" in name
        ]
        self.assertEqual(len(recovery_files), 1)
        recovery_path = os.path.join(
            self.temp_directory.name,
            recovery_files[0],
        )
        self.assertIsNone(self.cell_value(recovery_path, 2, 1))


class ExcelLibraryLifecycleTests(unittest.TestCase):
    def test_close_current_selects_next_workbook_after_close_failure(self):
        library = ExcelLibrary()
        first = Mock()
        first.close.side_effect = OSError("close failed")
        second = Mock()
        library._cache = {"first": first, "second": second}
        library._current_id = "first"

        with self.assertRaisesRegex(RuntimeError, "current Excel workbook"):
            library.close_current_excel_document()

        self.assertEqual(library._current_id, "second")
        self.assertEqual(library._cache, {"second": second})

    def test_close_all_attempts_every_workbook(self):
        library = ExcelLibrary()
        first = Mock()
        first.close.side_effect = OSError("close failed")
        second = Mock()
        library._cache = {"first": first, "second": second}
        library._current_id = "first"

        with self.assertRaisesRegex(RuntimeError, "1 Excel workbook"):
            library.close_all_excel_documents()

        first.close.assert_called_once_with()
        second.close.assert_called_once_with()
        self.assertEqual(library._cache, {})
        self.assertIsNone(library._current_id)


class ExecutionReportingTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_directory.cleanup)
        self.reporting = ExecutionReporting()
        self.manifest_path = self.reporting.initialize_execution_reporting(
            self.temp_directory.name,
            "Enterprise Batch 1",
            "worker-2",
        )

    def events(self):
        with open(self.manifest_path, encoding="utf-8") as stream:
            return [json.loads(line) for line in stream if line.strip()]

    def test_every_declared_failure_code_can_be_recorded(self):
        for index, failure_code in enumerate(sorted(FAILURE_CODES)):
            content_id = f"069{index:015d}"
            self.reporting.start_document_attempt(content_id)
            record = self.reporting.record_document_failure(
                content_id,
                failure_code,
                f"mapped {failure_code}",
                retryable=False,
            )
            self.assertEqual(record["FailureCode"], failure_code)

    def test_unknown_failure_code_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "Unknown failure code"):
            self.reporting.record_document_failure("069bad", "NOT_A_CODE", "bad")

    def test_attempt_count_includes_initial_attempt_and_retries(self):
        content_id = "069AAAAAAAAAAAAY55"
        for expected_attempt in (1, 2, 3):
            attempt = self.reporting.start_document_attempt(content_id)
            record = self.reporting.record_document_failure(
                content_id,
                "DOWNLOAD_APPEAR_TIMEOUT",
                "timed out",
            )
            self.assertEqual(attempt, expected_attempt)
            self.assertEqual(record["AttemptCount"], expected_attempt)

    def test_successful_retry_removes_final_failure(self):
        content_id = "069AAAAAAAAAAAAY55"
        self.reporting.start_document_attempt(content_id)
        self.reporting.record_document_failure(
            content_id, "DOWNLOAD_APPEAR_TIMEOUT", "first failure"
        )
        self.reporting.start_document_attempt(content_id)
        self.reporting.record_document_success(content_id, 25, "file.bin")

        self.assertEqual(self.reporting.get_failure_records([content_id]), [])
        self.assertEqual(self.events()[-1]["AttemptCount"], 2)
        self.assertNotIn(content_id, self.reporting._started)

    def test_retry_preserves_metadata_when_optional_values_are_omitted(self):
        content_id = "069AAAAAAAAAAAAY55"
        self.reporting.start_document_attempt(
            content_id, "068version", "Title", "file.pdf", "C:/file.pdf", 12, 2
        )
        self.reporting.record_document_failure(
            content_id, "DOWNLOAD_APPEAR_TIMEOUT", "first failure"
        )
        self.reporting.start_document_attempt(content_id)
        self.reporting.record_document_success(content_id, 12)

        event = self.events()[-1]
        self.assertEqual(event["ContentVersionId"], "068version")
        self.assertEqual(event["OriginalTitle"], "Title")
        self.assertEqual(event["ExpectedSize"], 12)
        self.assertEqual(event["ContentDocumentLinkCount"], 2)

    def test_manifest_io_failure_does_not_reclassify_committed_success(self):
        content_id = "069AAAAAAAAAAAAY55"
        self.reporting.start_document_attempt(content_id)
        with patch("builtins.open", side_effect=OSError("disk full")):
            record = self.reporting.record_document_success(content_id, 25)

        self.assertEqual(record["ActualSize"], 25)
        self.assertEqual(self.reporting.get_failure_records([content_id]), [])
        self.assertIn("disk full", self.reporting.get_execution_reporting_error())

    def test_redacts_tokens_and_frontdoor_session_ids(self):
        secret = "00Dxx0000000001!AQ0AQ-secret-token"
        text = (
            f"Authorization: Bearer {secret}; "
            f'{{"accessToken":"{secret}"}} '
            f"https://example.my.salesforce.com/secur/frontdoor.jsp?sid={secret}"
        )
        redacted = self.reporting.redact_sensitive_text(text)

        self.assertNotIn(secret, redacted)
        self.assertGreaterEqual(redacted.count("[REDACTED]"), 3)

    def test_spreadsheet_formula_prefix_is_escaped(self):
        self.assertEqual(
            self.reporting.sanitize_spreadsheet_cell('=HYPERLINK("bad")'),
            '\'=HYPERLINK("bad")',
        )

    def test_manifest_contains_ordered_terminal_events(self):
        content_id = "069AAAAAAAAAAAAY55"
        self.reporting.start_document_attempt(
            content_id, "068version", "Title", "file.pdf", "C:/file.pdf", 12, 2
        )
        self.reporting.record_document_success(content_id, 12, "C:/file.pdf")
        self.reporting.complete_execution_reporting(1, 0, 1)

        events = self.events()
        self.assertEqual(
            [event["EventType"] for event in events],
            [
                "EXECUTION_STARTED",
                "DOCUMENT_ATTEMPT_STARTED",
                "DOCUMENT_SUCCEEDED",
                "EXECUTION_COMPLETED",
            ],
        )
        self.assertTrue(all(event["SchemaVersion"] == 1 for event in events))
        self.assertTrue(all(event["WorkerId"] == "worker-2" for event in events))

    def test_detects_rest_and_browser_session_expiration(self):
        self.assertTrue(self.reporting.is_salesforce_auth_failure(401, ""))
        self.assertTrue(
            self.reporting.is_salesforce_auth_failure(
                403, '[{"errorCode":"INVALID_SESSION_ID"}]'
            )
        )
        self.assertTrue(
            self.reporting.is_salesforce_login_url(
                "https://acme.my.salesforce.com/secur/login_portal.jsp"
            )
        )
        self.assertFalse(
            self.reporting.is_salesforce_login_url(
                "https://acme.my.salesforce.com/lightning/page/home"
            )
        )

    def test_detects_html_download_response(self):
        path = os.path.join(self.temp_directory.name, "download.pdf")
        with open(path, "wb") as stream:
            stream.write(b"<!doctype html><html>Salesforce login</html>")
        self.assertTrue(self.reporting.file_looks_like_html(path))

    def test_does_not_treat_a_legitimate_html_document_as_expired_session(self):
        path = os.path.join(self.temp_directory.name, "customer.html")
        with open(path, "wb") as stream:
            stream.write(b"<!doctype html><html><body>Quarterly report</body></html>")
        self.assertFalse(self.reporting.file_looks_like_html(path))

    def test_robot_workflow_references_every_declared_failure_code(self):
        repository = Path(__file__).resolve().parents[2]
        resources = "\n".join(
            path.read_text(encoding="utf-8")
            for path in (repository / "src" / "robot" / "resources").glob("*.robot")
        )
        for failure_code in FAILURE_CODES:
            with self.subTest(failure_code=failure_code):
                self.assertIn(failure_code, resources)

    def test_download_failure_messages_are_mapped_to_expected_codes(self):
        repository = Path(__file__).resolve().parents[2]
        source = (
            repository / "src/robot/resources/download_operations.robot"
        ).read_text(encoding="utf-8")
        mappings = {
            "DOWNLOAD_NAVIGATION_FAILED": "Browser navigation to download URL failed",
            "DOWNLOAD_APPEAR_TIMEOUT": "Download file did not appear within timeout",
            "DOWNLOAD_COMPLETION_TIMEOUT": "Download did not complete within timeout",
            "MULTIPLE_FILES_NO_SIZE_MATCH": "No downloaded file matched expected file size",
            "CONTENT_SIZE_MISMATCH": "Downloaded file size does not match expected ContentSize",
            "FILE_NOT_STABLE": "Downloaded file size did not stabilize",
            "FILE_MOVE_FAILED": "Downloaded file remained locked",
            "WORKBOOK_TRANSACTION_FAILED": "Migration workbook update failed",
            "FINAL_FILE_VALIDATION_FAILED": "Moved file missing or final file size validation failed",
        }
        for failure_code, message in mappings.items():
            with self.subTest(failure_code=failure_code):
                self.assertRegex(
                    source,
                    rf"{re.escape(failure_code)}[\s\S]{{0,250}}{re.escape(message)}",
                )


class SalesforceSupportJsonTests(unittest.TestCase):
    def setUp(self):
        self.support = SalesforceSupport()

    def test_try_parse_returns_false_for_empty_output(self):
        parsed, value = self.support.try_parse_first_json_value("")

        self.assertFalse(parsed)
        self.assertIsNone(value)

    def test_try_parse_returns_false_for_invalid_output(self):
        parsed, value = self.support.try_parse_first_json_value(
            "Salesforce CLI warning without JSON"
        )

        self.assertFalse(parsed)
        self.assertIsNone(value)

    def test_try_parse_returns_first_valid_json_value(self):
        parsed, value = self.support.try_parse_first_json_value(
            'Warning\n{"status": 0}\nTrailing text'
        )

        self.assertTrue(parsed)
        self.assertEqual(value, {"status": 0})


class SalesforceSupportIdTests(unittest.TestCase):
    def setUp(self):
        self.support = SalesforceSupport()

    def test_converts_valid_15_character_id_to_18_characters(self):
        self.assertEqual(
            self.support.canonicalize_content_document_id("069AAAAAAAAAAAA"),
            "069AAAAAAAAAAAAY55",
        )

    def test_preserves_existing_18_character_id(self):
        content_id = "069AAAAAAAAAAAAY55"
        self.assertEqual(
            self.support.canonicalize_content_document_id(content_id),
            content_id,
        )

    def test_preserves_invalid_value_for_workflow_validation(self):
        self.assertEqual(
            self.support.canonicalize_content_document_id(" invalid "),
            "invalid",
        )


class SalesforceSupportPathTests(unittest.TestCase):
    def test_complete_filename_respects_destination_path_budget(self):
        support = SalesforceSupport()
        directory = os.path.join(
            tempfile.gettempdir(),
            "destination",
            "069000000000001",
        )

        filename = support.sanitize_local_filename_for_directory(
            f"{'x' * 300}.pdf",
            directory,
            max_filename_length=180,
            max_path_length=120,
        )

        complete_path = os.path.join(os.path.abspath(directory), filename)
        self.assertLessEqual(len(complete_path), 120)
        self.assertTrue(filename.endswith(".pdf"))

    def test_rejects_directory_without_filename_budget(self):
        support = SalesforceSupport()
        with self.assertRaisesRegex(ValueError, "no room"):
            support.sanitize_local_filename_for_directory(
                "file.txt",
                "x" * 30,
                max_path_length=10,
            )


class SalesforceSupportDownloadWaitTests(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_directory.cleanup)
        self.support = SalesforceSupport()

    def test_accepts_completed_file_without_transient_failures(self):
        completed_path = os.path.join(self.temp_directory.name, "complete.pdf")
        with open(completed_path, "wb") as completed_file:
            completed_file.write(b"complete")

        result = self.support.wait_for_completed_download(
            self.temp_directory.name,
            0,
            [".crdownload", ".tmp", ".part"],
        )

        self.assertTrue(result)

    def test_times_out_when_only_temporary_file_exists(self):
        temporary_path = os.path.join(self.temp_directory.name, "pending.tmp")
        with open(temporary_path, "wb") as temporary_file:
            temporary_file.write(b"pending")

        with self.assertRaisesRegex(TimeoutError, "No completed download"):
            self.support.wait_for_completed_download(
                self.temp_directory.name,
                0,
                [".crdownload", ".tmp", ".part"],
            )
